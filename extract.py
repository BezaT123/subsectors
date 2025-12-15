import openpyxl
import json
from datetime import datetime
import os
import argparse
import re

def find_sheet_by_variations(workbook, primary_names, fallback_names=None):
    """
    Find a sheet by trying various name variations (case-insensitive, spaces, underscores).
    
    Args:
        workbook: The openpyxl workbook object
        primary_names: List of primary names to search for (e.g., ['i_Setup', 'i Setup', 'Setup'])
        fallback_names: Optional list of fallback names if primary names not found (e.g., ['summary', 'Summary'])
    
    Returns:
        str: The actual sheet name found, or None if not found
    """
    all_names = primary_names.copy()
    if fallback_names:
        all_names.extend(fallback_names)
    
    # Normalize function: remove spaces, underscores, convert to lowercase
    def normalize(name):
        return re.sub(r'[\s_]+', '', name.lower())
    
    # First try exact matches (case-insensitive)
    for name in all_names:
        for sheet_name in workbook.sheetnames:
            if sheet_name.lower() == name.lower():
                return sheet_name
    
    # Then try normalized matches (ignoring spaces and underscores)
    normalized_targets = {normalize(name): name for name in all_names}
    for sheet_name in workbook.sheetnames:
        normalized_sheet = normalize(sheet_name)
        if normalized_sheet in normalized_targets:
            return sheet_name
    
    # Try partial matches (contains the key word)
    for name in all_names:
        key_word = normalize(name)
        for sheet_name in workbook.sheetnames:
            if key_word in normalize(sheet_name) or normalize(sheet_name) in key_word:
                return sheet_name
    
    return None

def extract_setup_data_to_json(excel_file_path, output_json_path=None):
    """
    Extract data from the i_Setup and i_COS sheets of an Excel file and convert to JSON format.
    
    Args:
        excel_file_path (str): Path to the Excel file (.xlsx or .xlsm)
        output_json_path (str, optional): Path for the output JSON file. 
                                        If None, creates a file with same name as Excel file.
    
    Returns:
        dict: The extracted data as a dictionary
    """
    
    try:
        # Load the Excel workbook
        workbook = openpyxl.load_workbook(excel_file_path, data_only=True)
        
        # Find i_Setup sheet with variations, fallback to 'summary'
        setup_sheet_name = find_sheet_by_variations(
            workbook,
            primary_names=['i_Setup', 'i Setup', 'Setup', 'i-Setup'],
            fallback_names=['summary', 'Summary', 'Summary Sheet']
        )
        
        if setup_sheet_name is None:
            available_sheets = ', '.join(workbook.sheetnames)
            raise ValueError(f"Sheet 'i_Setup' (or variations) or 'summary' not found in the workbook. Available sheets: {available_sheets}")
        
        print(f"Using sheet: '{setup_sheet_name}' for i_Setup data")
        
        # Find i_COS sheet with variations
        cos_sheet_name = find_sheet_by_variations(
            workbook,
            primary_names=['i_COS', 'i COS', 'COS', 'i-COS', 'Cost of Sales']
        )
        
        if cos_sheet_name is None:
            available_sheets = ', '.join(workbook.sheetnames)
            raise ValueError(f"Sheet 'i_COS' (or variations) not found in the workbook. Available sheets: {available_sheets}")
        
        print(f"Using sheet: '{cos_sheet_name}' for i_COS data")
        
        # Check for 'info' sheet and extract
        info_data = {}
        info_sheet_name = find_sheet_by_variations(
            workbook,
            primary_names=['info', 'Info', 'Information']
        )
        if info_sheet_name:
            info_data = extract_info_data(workbook[info_sheet_name])
        else:
            print("Warning: Sheet 'info' not found, skipping benchmark data extraction.")
        
        # Check for 'financials' sheet and extract
        financials_data = {}
        financials_sheet_name = find_sheet_by_variations(
            workbook,
            primary_names=['financials', 'Financials', 'Financial', 'Financial Statements', 'FS']
        )
        if financials_sheet_name:
            print(f"Using sheet: '{financials_sheet_name}' for financials data")
            financials_data = extract_financials_data(workbook[financials_sheet_name])
        else:
            print("Warning: Sheet 'financials' not found, skipping financials data extraction.")
        
        # Extract i_Setup data
        setup_data = extract_i_setup_data(workbook[setup_sheet_name])
        
        # Extract i_COS data
        cos_data = extract_i_cos_data(workbook[cos_sheet_name])
        
        # Create the final JSON structure
        result = {
            "extractedAt": datetime.now().isoformat(),
            "sourceFile": os.path.basename(excel_file_path),
            "i_Setup": {
                "totalFields": len(setup_data),
                "fieldsWithSubTables": len([f for f in setup_data.values() if f["hasSubTable"]]),
                "totalSubTableItems": sum(len(f["subTableData"]) for f in setup_data.values()),
                "fields": setup_data
            },
            "i_COS": {
                "totalProducts": len(cos_data),
                "products": cos_data
            },
            "info": info_data,
            "financials": financials_data
        }
        
        # Generate output file path if not provided
        if output_json_path is None:
            base_name = os.path.splitext(excel_file_path)[0]
            output_json_path = f"{base_name}.json"
        
        # Save to JSON file
        with open(output_json_path, 'w', encoding='utf-8') as json_file:
            json.dump(result, json_file, indent=2, ensure_ascii=False)
        
        print(f"Extraction completed successfully!")
        print(f"i_Setup - Total fields: {len(setup_data)}")
        print(f"i_COS - Total products: {len(cos_data)}")
        if financials_data:
            total_line_items = sum(len(items) for items in financials_data.values())
            categories = list(financials_data.keys())
            print(f"Financials - Total line items: {total_line_items}, Categories: {len(categories)} ({', '.join(categories[:5])}{'...' if len(categories) > 5 else ''})")
        print(f"JSON file saved as: {output_json_path}")
        
        return result
        
    except FileNotFoundError:
        print(f"Error: Excel file '{excel_file_path}' not found.")
        return None
    except Exception as e:
        print(f"Error occurred: {str(e)}")
        return None

def extract_i_setup_data(worksheet):
    """Extract data from i_Setup sheet"""
    setup_data = {}
    
    # Extract basic field definitions
    for row_num in range(1, worksheet.max_row + 1):
        field_number_cell = worksheet.cell(row=row_num, column=7)  # Column G
        field_name_cell = worksheet.cell(row=row_num, column=8)    # Column H
        field_type_cell = worksheet.cell(row=row_num, column=9)    # Column I
        field_value_cell = worksheet.cell(row=row_num, column=13)  # Column M
        
        if (field_number_cell.value is not None and 
            isinstance(field_number_cell.value, (int, float)) and
            field_name_cell.value is not None and
            isinstance(field_name_cell.value, str)):
            
            field_number = int(field_number_cell.value)
            field_name = str(field_name_cell.value).strip()
            field_type = str(field_type_cell.value).strip() if field_type_cell.value else ""
            field_value = field_value_cell.value if field_value_cell.value is not None else ""
            
            if not isinstance(field_value, str):
                field_value = str(field_value) if field_value != "" else ""
            
            setup_data[field_name] = {
                "fieldNumber": field_number,
                "fieldType": field_type,
                "value": field_value,
                "hasSubTable": False,
                "subTableData": []
            }
    
    # Handle Industry Details special case
    if "Industry Details" in setup_data:
        industry_data = extract_industry_details_subtable(worksheet)
        if industry_data:
            setup_data["Industry Details"]["hasSubTable"] = True
            setup_data["Industry Details"]["subTableData"] = industry_data

    # Extract sub-table data for complex fields
    field_sections = [
        {"name": "Revenue Streams", "start_row": 44, "end_row": 58},
        {"name": "Cost of Sales", "start_row": 59, "end_row": 77},
        {"name": "Operating Costs", "start_row": 78, "end_row": 96},
        {"name": "Financing Costs", "start_row": 97, "end_row": 105},
        {"name": "Capital Investment", "start_row": 106, "end_row": 116},
        {"name": "Borrowing Details", "start_row": 117, "end_row": 129},
        {"name": "Credit Scoring Details", "start_row": 130, "end_row": 136},
        {"name": "Cash Flow Details", "start_row": 137, "end_row": 146},
        {"name": "Collateral & Other", "start_row": 147, "end_row": 158},
        {"name": "Recommendations", "start_row": 159, "end_row": 170}
    ]
    
    for section in field_sections:
        field_name = section["name"]
        start_row = section["start_row"]
        end_row = section["end_row"]
        
        if field_name in setup_data:
            sub_table_data = []
            header_found = False
            
            for row_num in range(start_row, min(end_row + 1, worksheet.max_row + 1)):
                # Check for sub-table header row
                name_header = worksheet.cell(row=row_num, column=13)  # Column M
                type_header = worksheet.cell(row=row_num, column=15)  # Column O
                
                if (name_header.value == "Name" and type_header.value == "Type"):
                    header_found = True
                    continue
                
                # Process data rows after header is found
                if header_found:
                    item_name_cell = worksheet.cell(row=row_num, column=8)   # Column H
                    item_type_cell = worksheet.cell(row=row_num, column=9)   # Column I
                    name_cell = worksheet.cell(row=row_num, column=13)       # Column M
                    type_cell = worksheet.cell(row=row_num, column=15)       # Column O
                    industry_cell = worksheet.cell(row=row_num, column=16)   # Column P
                    sub1_cell = worksheet.cell(row=row_num, column=17)       # Column Q
                    sub2_cell = worksheet.cell(row=row_num, column=18)       # Column R
                    sub3_cell = worksheet.cell(row=row_num, column=19)       # Column S
                    
                    # Check if this row contains sub-table item data
                    if (item_name_cell.value and 
                        isinstance(item_name_cell.value, str) and 
                        ('Category' in item_name_cell.value or 'Steam' in item_name_cell.value)):
                        
                        sub_item = {
                            "itemName": str(item_name_cell.value).strip() if item_name_cell.value else "",
                            "itemType": str(item_type_cell.value).strip() if item_type_cell.value else "",
                            "name": str(name_cell.value).strip() if name_cell.value else "",
                            "type": str(type_cell.value).strip() if type_cell.value else "",
                            "industry": str(industry_cell.value).strip() if industry_cell.value else "",
                            "sub1": str(sub1_cell.value).strip() if sub1_cell.value else "",
                            "sub2": str(sub2_cell.value).strip() if sub2_cell.value else "",
                            "sub3": str(sub3_cell.value).strip() if sub3_cell.value else ""
                        }
                        
                        # Only add if there's meaningful data
                        if sub_item["name"] or sub_item["type"] or sub_item["sub1"] or sub_item["sub2"] or sub_item["sub3"]:
                            sub_table_data.append(sub_item)
            
            # Update field data with sub-table information
            if sub_table_data:
                setup_data[field_name]["hasSubTable"] = True
                setup_data[field_name]["subTableData"] = sub_table_data
    
    return setup_data

def extract_industry_details_subtable(worksheet):
    """Extract the special Industry Details sub-table"""
    industry_data = []
    
    # Industry Details appears around rows 19-26 based on the field structure
    for row_num in range(19, 27):
        field_label_cell = worksheet.cell(row=row_num, column=8)   # Column H
        field_type_cell = worksheet.cell(row=row_num, column=9)    # Column I
        field_value_cell = worksheet.cell(row=row_num, column=13)  # Column M
        
        if (field_label_cell.value and 
            isinstance(field_label_cell.value, str) and 
            field_label_cell.value in ["Industry Type", "Primary Industry", "Secondary Industry", "Benchmarking Business Sector"]):
            
            industry_item = {
                "fieldLabel": str(field_label_cell.value).strip(),
                "fieldType": str(field_type_cell.value).strip() if field_type_cell.value else "",
                "value": str(field_value_cell.value).strip() if field_value_cell.value else ""
            }
            
            industry_data.append(industry_item)
    
    return industry_data

def extract_i_cos_data(worksheet):
    """
    Extract product data from i_COS sheet with enhanced information extraction.
    This now captures both product names and their cost of sales categories more comprehensively.
    """
    cos_data = []
    
    for row_num in range(1, worksheet.max_row + 1):
        product_name_cell = worksheet.cell(row=row_num, column=8)   # Column H
        cos_category_cell = worksheet.cell(row=row_num, column=9)   # Column I
        
        # Check if both cells have meaningful data
        if (product_name_cell.value and 
            cos_category_cell.value and
            isinstance(product_name_cell.value, str) and
            isinstance(cos_category_cell.value, str)):
            
            product_name = str(product_name_cell.value).strip()
            cos_category = str(cos_category_cell.value).strip()
            
            # Skip header rows or empty data, but be more inclusive of actual data
            if (product_name and cos_category and 
                product_name not in ["Product Name", ""] and
                cos_category not in ["Cost of Sales Category", ""]):
                
                cos_data.append({
                    "productName": product_name,
                    "costOfSalesCategory": cos_category
                })
    
    # Remove any duplicate entries while preserving order
    seen = set()
    unique_cos_data = []
    for item in cos_data:
        # Create a unique identifier for each item
        identifier = (item["productName"], item["costOfSalesCategory"])
        if identifier not in seen:
            seen.add(identifier)
            unique_cos_data.append(item)
    
    return unique_cos_data

def extract_info_data(worksheet):
    """
    Extract all benchmarkable metrics from the 'info' sheet generically.
    Extracts label-value pairs where label is in Column A and value is in Column B.
    Returns a flat dictionary with normalized keys (snake_case) and original values.
    Also includes a '_labels' mapping to preserve original label names.
    """
    info_data = {}
    label_mapping = {}  # Maps normalized keys to original labels
    
    # Normalize label to snake_case key
    def normalize_key(label):
        """Convert label to snake_case key"""
        # Remove special characters, replace spaces/underscores with single underscore
        normalized = re.sub(r'[^\w\s]', '', label)
        normalized = re.sub(r'[\s_]+', '_', normalized)
        normalized = normalized.lower().strip('_')
        return normalized
    
    # Search for label-value pairs across the sheet
    # Typically structured as: label in Column A, value in Column B
    for row_num in range(1, worksheet.max_row + 1):
        label_cell = worksheet.cell(row=row_num, column=1)  # Column A
        value_cell = worksheet.cell(row=row_num, column=2)   # Column B
        
        # Check if label exists and is meaningful
        if label_cell.value is not None:
            label = str(label_cell.value).strip()
            
            # Skip empty labels, headers, or common non-data rows
            if (label and 
                label.lower() not in ['', 'label', 'metric', 'name', 'description', 'info'] and
                not label.startswith('#')):
                
                # Get the value (can be string, number, or None)
                value = value_cell.value
                
                # Only add if there's a meaningful value
                if value is not None:
                    # Normalize the label to create a consistent key
                    key = normalize_key(label)
                    
                    # Store the value with normalized key, and preserve original label
                    if key:  # Only add if we got a valid key
                        info_data[key] = value
                        label_mapping[key] = label
    
    # Add label mapping as metadata if we have any data
    if label_mapping:
        info_data['_labels'] = label_mapping
    
    return info_data

def extract_financials_data(worksheet):
    """
    Extract subcategories and sub-subcategories from the financials sheet grouped by category.
    Assumes:
    - Categories are in Column A (sub1) or Column E
    - Subcategories are in Column F
    - Sub-subcategories are in Column G (or comments)
    Returns a hierarchical structure: category -> subcategory -> [sub-subcategories]
    Expected categories: revenue, cost_of_sale, opex, financing_cost, capex
    """
    financials_data = {}
    
    # Expected category keys
    EXPECTED_CATEGORIES = ['revenue', 'cost_of_sale', 'opex', 'financing_cost', 'capex']
    
    # Map category names to normalized keys
    def normalize_category(category):
        """Convert category to normalized key (lowercase, underscores).
        Only returns one of the expected categories: revenue, cost_of_sale, opex, financing_cost, capex
        """
        if not category:
            return None
        
        category_str = str(category).strip()
        
        # Normalize the string
        normalized = re.sub(r'[^\w\s]', '', category_str)
        normalized = re.sub(r'[\s_]+', '_', normalized)
        normalized = normalized.lower().strip('_')
        
        # Map common variations to expected keys - prioritize cost_of_sale vs opex
        category_mapping = {
            # Revenue
            'revenue': 'revenue',
            'revenues': 'revenue',
            'income': 'revenue',
            'sales': 'revenue',
            
            # Cost of Sale (direct costs related to products/services sold)
            'cost_of_sale': 'cost_of_sale',
            'cost_of_sales': 'cost_of_sale',
            'cogs': 'cost_of_sale',
            'cost_of_goods_sold': 'cost_of_sale',
            'cost_of_goods': 'cost_of_sale',
            'direct_cost': 'cost_of_sale',
            'direct_costs': 'cost_of_sale',
            'cost_of_revenue': 'cost_of_sale',
            
            # OPEX (operating expenses - indirect costs)
            'opex': 'opex',
            'operating_expenses': 'opex',
            'operating_expense': 'opex',
            'operating_costs': 'opex',
            'operating_cost': 'opex',
            'operating_expenditure': 'opex',
            'operating_expenditures': 'opex',
            'expenses': 'opex',
            'expenditure': 'opex',
            'expenditures': 'opex',
            
            # Financing Cost
            'financing_cost': 'financing_cost',
            'financing_costs': 'financing_cost',
            'finance_cost': 'financing_cost',
            'finance_costs': 'financing_cost',
            'interest_expense': 'financing_cost',
            'interest_expenses': 'financing_cost',
            'interest_cost': 'financing_cost',
            'interest_costs': 'financing_cost',
            'financial_cost': 'financing_cost',
            'financial_costs': 'financing_cost',
            
            # CAPEX (capital expenditure)
            'capex': 'capex',
            'capital_expenditure': 'capex',
            'capital_expenditures': 'capex',
            'capital_investment': 'capex',
            'capital_investments': 'capex',
            'capital_expense': 'capex',
            'capital_expenses': 'capex',
        }
        
        # First check exact match
        if normalized in category_mapping:
            return category_mapping[normalized]
        
        # Check for partial matches (be careful with order - check more specific first)
        # Check for cost_of_sale indicators first (before opex)
        if any(term in normalized for term in ['cost_of_sale', 'cost_of_sales', 'cogs', 'cost_of_goods', 'direct_cost']):
            return 'cost_of_sale'
        
        # Check for opex indicators
        if any(term in normalized for term in ['opex', 'operating_expense', 'operating_cost', 'operating_expenditure']):
            return 'opex'
        
        # Check for revenue indicators
        if any(term in normalized for term in ['revenue', 'income', 'sales']):
            return 'revenue'
        
        # Check for financing cost indicators
        if any(term in normalized for term in ['financing', 'finance', 'interest']):
            return 'financing_cost'
        
        # Check for capex indicators
        if any(term in normalized for term in ['capex', 'capital_expenditure', 'capital_investment', 'capital_expense']):
            return 'capex'
        
        # If no match found, return None (don't include unknown categories)
        return None
    
    # Extract line items starting from row 1
    start_row = 1
    line_items_by_category = {}
    
    for row_num in range(start_row, worksheet.max_row + 1):
        # Get category from Column E or Column A (sub1)
        # Try both and use the more specific one
        category = None
        category_raw = None
        category_from_e = None
        category_from_a = None
        
        # Try Column A (sub1) first - this is usually more specific (e.g., "Cost of Sale", "OPEX")
        # Column E might have generic terms (e.g., "Expenditure") for everything
        sub1_cell = worksheet.cell(row=row_num, column=1)  # Column A for sub1
        if sub1_cell.value:
            sub1_value = str(sub1_cell.value).strip()
            if sub1_value.lower() not in ["", "sub1", "line item", "account", "description", "name"]:
                category_from_a = normalize_category(sub1_value)
        
        # Fallback to Column E if sub1 didn't give us a valid category
        if not category_from_a or category_from_a not in EXPECTED_CATEGORIES:
            category_cell_e = worksheet.cell(row=row_num, column=5)  # Column E
            if category_cell_e.value:
                category_raw = str(category_cell_e.value).strip()
                # Skip if it looks like a header
                if category_raw.lower() not in ["", "sub1", "category", "type", "comments", "sub3"]:
                    category_from_e = normalize_category(category_raw)
        
        # Prefer Column A (sub1) over Column E since sub1 is usually more specific
        # If both exist and are valid, use sub1 (Column A)
        if category_from_a and category_from_a in EXPECTED_CATEGORIES:
            category = category_from_a
        elif category_from_e and category_from_e in EXPECTED_CATEGORIES:
            category = category_from_e
        
        # Skip if no valid category found or if it's not one of the expected categories
        if not category or category not in EXPECTED_CATEGORIES:
            continue
        
        # Get subcategory from Column F
        subcategory = None
        subcategory_cell_f = worksheet.cell(row=row_num, column=6)  # Column F
        if subcategory_cell_f.value:
            subcategory_value = str(subcategory_cell_f.value).strip()
            # Skip if it's a header or looks like a number
            if (subcategory_value.lower() not in ["", "sub3", "comments", "type", "category", "subcategory"] and
                not subcategory_value.replace(',', '').replace('.', '').replace('-', '').strip().isdigit()):
                subcategory = subcategory_value
        
        # Get sub-subcategory from Column G (most granular)
        sub_subcategory = None
        sub_subcategory_cell_g = worksheet.cell(row=row_num, column=7)  # Column G
        if sub_subcategory_cell_g.value:
            sub_subcategory_value = str(sub_subcategory_cell_g.value).strip()
            # Skip if it's a header or looks like a number
            if (sub_subcategory_value.lower() not in ["", "sub3", "comments", "type", "category", "subcategory"] and
                not sub_subcategory_value.replace(',', '').replace('.', '').replace('-', '').strip().isdigit()):
                sub_subcategory = sub_subcategory_value
        
        # Skip if we don't have at least a subcategory or sub-subcategory
        if not subcategory and not sub_subcategory:
            continue
        
        # Initialize category structure if needed
        if category not in line_items_by_category:
            line_items_by_category[category] = {}
        
        # If we have a subcategory
        if subcategory:
            if subcategory not in line_items_by_category[category]:
                line_items_by_category[category][subcategory] = []
            
            # If we also have a sub-subcategory, add it to the subcategory's list
            if sub_subcategory and sub_subcategory not in line_items_by_category[category][subcategory]:
                line_items_by_category[category][subcategory].append(sub_subcategory)
        # If we only have sub-subcategory (no subcategory), create a default grouping
        elif sub_subcategory:
            if "_other" not in line_items_by_category[category]:
                line_items_by_category[category]["_other"] = []
            if sub_subcategory not in line_items_by_category[category]["_other"]:
                line_items_by_category[category]["_other"].append(sub_subcategory)
    
    # Convert to the desired format
    financials_data = line_items_by_category
    
    return financials_data

def process_single_file(excel_file_path):
    """Process a single Excel file"""
    print(f"\nProcessing: {excel_file_path}")
    
    # Extract data and write per-file JSON using the Excel base name
    extracted_data = extract_setup_data_to_json(excel_file_path)
    
    if extracted_data:
        print(f"\nExtraction complete. Data saved to JSON file.")
        
        # Print summary of Cost of Sales categories found for verification
        cos_products = extracted_data.get('i_COS', {}).get('products', [])
        cos_categories = list(set(item.get('costOfSalesCategory', '') for item in cos_products))
        cos_categories = [cat for cat in cos_categories if cat]  # Remove empty strings
        
        if cos_categories:
            print(f"\nCost of Sales Categories found:")
            for category in cos_categories:
                print(f"  - {category}")
        
        # Print products found for verification
        product_names = [item.get('productName', '') for item in cos_products if item.get('productName', '')]
        if product_names:
            print(f"\nProducts found ({len(product_names)}):")
            for i, product in enumerate(product_names[:10]):  # Show first 10
                print(f"  - {product}")
            if len(product_names) > 10:
                print(f"  ... and {len(product_names) - 10} more products")
    
    return extracted_data


def process_batch_directory(directory_path):
    """Process all Excel files in a directory"""
    allowed_exts = (".xlsm", ".xlsx")
    
    if not os.path.isdir(directory_path):
        print(f"Error: '{directory_path}' is not a valid directory.")
        return
    
    excel_files = []
    for filename in sorted(os.listdir(directory_path)):
        if filename.startswith('.'):  # skip hidden files
            continue
        if not filename.lower().endswith(allowed_exts):
            continue
        if filename == 'subsectors-example.xlsx':  # skip reference file
            continue
        
        excel_file = os.path.join(directory_path, filename)
        excel_files.append(excel_file)
    
    if not excel_files:
        print(f"No Excel files found in directory: {directory_path}")
        return
    
    print(f"Found {len(excel_files)} Excel file(s) to process in {directory_path}\n")
    
    successful = 0
    failed = 0
    
    for excel_file in excel_files:
        try:
            process_single_file(excel_file)
            successful += 1
        except Exception as e:
            print(f"Error processing {excel_file}: {e}")
            failed += 1
    
    print(f"\n{'='*60}")
    print(f"Batch processing complete!")
    print(f"Successfully processed: {successful} file(s)")
    if failed > 0:
        print(f"Failed: {failed} file(s)")
    print(f"{'='*60}")


# Example usage
if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description='Extract data from Excel files (.xlsx/.xlsm) to JSON format.',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Process a single file
  python extract.py --file "Financials for Shalom Edited.xlsm"
  
  # Process all Excel files in a directory
  python extract.py --batch ./financial_analysis
  
  # Process all Excel files in current directory
  python extract.py --batch .
        """
    )
    
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument(
        '--file', '-f',
        type=str,
        help='Path to a specific Excel file to process'
    )
    group.add_argument(
        '--batch', '-b',
        type=str,
        help='Path to directory containing Excel files to process in batch'
    )
    
    args = parser.parse_args()
    
    if args.file:
        # Single file mode
        if not os.path.isfile(args.file):
            print(f"Error: File '{args.file}' not found.")
            exit(1)
        process_single_file(args.file)
    
    elif args.batch:
        # Batch directory mode
        process_batch_directory(args.batch)